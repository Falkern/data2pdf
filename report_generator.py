import pandas as pd
from sqlalchemy import create_engine
from fpdf import FPDF
import argparse
import logging
import openpyxl
import time
from tqdm import tqdm
from datetime import datetime

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def read_csv(file_path, chunksize=1000):
    if not file_path.endswith('.csv'):
        logging.error("Invalid file format. Please provide a CSV file.")
        return None
    
    try:
        logging.info("Fetching data from CSV...")
        df = pd.concat([chunk for chunk in tqdm(pd.read_csv(file_path, chunksize=chunksize), desc="Reading CSV")])
        logging.info(f"Successfully read data from {file_path}")
        return df
    except Exception as e:
        logging.error(f"Error reading CSV file: {e}")
        return None

def read_excel(file_path):
    if not file_path.endswith('.xlsx'):
        logging.error("Invalid file format. Please provide an Excel (.xlsx) file.")
        return None
    
    try:
        logging.info("Fetching data from Excel...")
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        data = [row for row in tqdm(sheet.iter_rows(values_only=True), desc="Reading Excel")]
        
        df = pd.DataFrame(data[1:], columns=data[0])
        logging.info(f"Successfully read data from {file_path}")
        return df
    except Exception as e:
        logging.error(f"Error reading Excel file: {e}")
        return None

def read_from_db(connection_string, query):
    try:
        logging.info("Connecting to the database and fetching data...")
        engine = create_engine(connection_string)
        df = pd.read_sql(query, engine)
        logging.info("Successfully fetched data from the database.")
        return df
    except Exception as e:
        logging.error(f"Error reading from database: {e}")
        return None

def generate_pdf(dataframe, output_path):
    if dataframe is None or dataframe.empty:
        logging.warning("No data to generate a PDF report.")
        return

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Arial", size=10)

    column_width = pdf.w / len(dataframe.columns) - 1

    for column in dataframe.columns:
        pdf.cell(column_width, 10, column, border=1)
    pdf.ln()

    for _, row in dataframe.iterrows():
        for value in row:
            pdf.cell(column_width, 10, str(value), border=1)
        pdf.ln()

    try:
        pdf.output(output_path)
        logging.info(f"PDF report saved as {output_path}")
    except Exception as e:
        logging.error(f"Error generating PDF: {e}")

def main():
    parser = argparse.ArgumentParser(description="Generate PDF report from CSV, Excel, or database")
    parser.add_argument("--csv", help="Path to the CSV file")
    parser.add_argument("--excel", help="Path to the Excel file")
    parser.add_argument("--db", help="Database connection string")
    parser.add_argument("--query", help="SQL query for database")
    parser.add_argument("--output", help="Output PDF file name")

    args = parser.parse_args()
    
    if not args.output:
        args.output = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

    start_time = time.time()

    if args.csv:
        df = read_csv(args.csv)
    elif args.excel:
        df = read_excel(args.excel)
    elif args.db and args.query:
        df = read_from_db(args.db, args.query)
    else:
        logging.error("Please provide either a CSV file path, an Excel file path, or a database connection string and a query.")
        return

    if df is not None:
        logging.info(f"Data Summary: {df.shape[0]} rows, {df.shape[1]} columns")
        generate_pdf(df, args.output)

    end_time = time.time()
    logging.info(f"Processing completed in {end_time - start_time:.2f} seconds")

if __name__ == "__main__":
    main()
